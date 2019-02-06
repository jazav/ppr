from openpyxl import load_workbook
import sys, getopt
import logging

WORK_BOOK_NAME = "счет общий балл и результат"
START_ROW = 11
FINAL_COL = 12


def get_files(argv):
    from_file = ''
    to_file = ''
    try:
        opts, args = getopt.getopt(argv, "hf:t:", ["from_file=", "to_file="])
    except getopt.GetoptError:
        print('ppr.py -f <from_file1<,from_file2>,<from_file3>...> -t <to_file>')
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print('test.py -f <from_file> -t <to_file>')
            sys.exit()
        elif opt in ("-f", "--from_files"):
            from_file = arg
        elif opt in ("-t", "--to_file"):
            to_file = arg
    return from_file, to_file


def get_from_items(from_file):
    wb = load_workbook(filename=from_file, read_only=True, data_only=True)
    # , data_only=True
    ws = wb[WORK_BOOK_NAME]

    from_items = dict()

    for row in range(START_ROW, ws.rows.gi_frame.f_locals['max_row'] + 1):
        if ws[row][5].value is not None:
            item = []
            for col in range(1, FINAL_COL):
                # last col is 11
                item.append(ws[row][col].value)

            from_items[ws[row][0].value] = item

    return from_items


def find_row_by_code(ws, code):
    for row in range(START_ROW, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == code:
            return row


def fill_to_file(to_file, final_file, from_items):
    # MUSTN'T change data_only to True
    wb = load_workbook(filename=to_file, data_only=False)
    ws = wb[WORK_BOOK_NAME]

    for key, value in from_items.items():
        logging.info('writing: %s, %s', key, value[1])
        row = find_row_by_code(ws, key)

        for i in range(len(value)):
            if value[i] != "":
                ws[row][i + 1].value = value[i]
    wb.save(final_file)

def main(argv):
    from_files, to_file = get_files(argv)
    from_file_list = from_files.split(",")
    new_file = ""
    for from_file in from_file_list:
        from_items = get_from_items(from_file)
        logging.info('from file: %s',from_file)
        logging.info('to file: %s',to_file)
        logging.info('total people: %s', len(from_items))
        logging.debug(from_items)
        if new_file == "":
            new_file = to_file
        else:
            new_file = "new- " + to_file
        fill_to_file(new_file, "new-" + to_file, from_items)

if __name__ == '__main__':
    logger = logging.getLogger('ppr_logger')
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s', filemode='w')
    logger.addHandler(logging.FileHandler('ppr.log'))
    logger.addHandler(logging.StreamHandler(sys.stdout))

    main(sys.argv[1:])
